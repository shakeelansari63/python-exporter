from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Color, Font, PatternFill
import re
from .cenums import TemplateName

max_row = 0
    
def export_to_xlsx(data, template, template_name: TemplateName):
    wb = Workbook()
    ws = wb.active
    ws.title = template_name 

    for section in template.keys():
        if section == "head":
            generate_static_section_export(template[section], wb, ws, data)
        elif section == "table":
            generate_table_section_export(template[section], wb, ws, data)
        elif section == "foot":
            generate_static_section_export(template[section], wb, ws, data)
        else:
            print(f"Invalid Section: {section}")

    wb.save('Output.xlsx')
            

def generate_static_section_export(section, workbook, worksheet, inp_data):
    global max_row
    for field in section:
        field_data = replace_var( field['data'] if 'data' in field.keys() else '', inp_data)
        position = replace_var( field['start'] if 'start' in field.keys() else [0, 0], inp_data)
        colspan = replace_var( field['colspan'] if 'colspan' in field.keys() else 1, inp_data)
        rowspan = replace_var(field['rowspan'] if 'rowspan' in field.keys() else 1, inp_data)
        align = replace_var( field['align'] if 'align' in field.keys() else 'left', inp_data)
        valign = replace_var( field['valign'] if 'valign' in field.keys() else 'center', inp_data)
        border = replace_var( field['border'] if 'border' in field.keys() else False, inp_data)
        bgcolor = replace_var( field['bgcolor'] if 'bgcolor' in field.keys() else None, inp_data)
        fgcolor = replace_var( field['fgcolor'] if 'fgcolor' in field.keys() else None, inp_data)
        bold = replace_var( field['bold'] if 'bold' in field.keys() else False, inp_data)

        bgcolor = bgcolor.lstrip('#') if bgcolor is not None else bgcolor
        fgcolor = fgcolor.lstrip('#') if fgcolor is not None else fgcolor

        if type(position[0]) == type(''):
            position[0] = eval(position[0])

        if type(position[1]) == type(''):
            position[1] = eval(position[1])

        if position[0] > max_row:
            max_row = position[0]

        if colspan > 1 or rowspan > 1:
            worksheet.merge_cells(start_row = position[0], 
                                    start_column = position[1], 
                                    end_row = position[0] + rowspan - 1, 
                                    end_column = position[1] + colspan - 1)

        cell = worksheet.cell(row = position[0], column = position[1])
        cell.value = field_data
        cell.alignment = Alignment(horizontal = align, vertical = valign)

        if bgcolor is not None:
            cell.fill = PatternFill(bgColor = bgcolor, fgColor = bgcolor, fill_type = "solid")

        if border:
            cell.border = Border(left = Side(border_style=None, color='000000', style='thin'),
                                     right = Side(border_style=None, color='000000', style='thin'),
                                     top = Side(border_style=None, color='000000', style='thin'),
                                     bottom = Side(border_style=None, color='000000', style='thin'))

        if bold or fgcolor is not None:
            cell.font = Font(bold = bold, color = fgcolor)

def generate_table_section_export(section, workbook, worksheet, inp_data):
    global max_row
    table_data = replace_var(section['data'] if 'data' in section.keys() else [], inp_data)
    position = replace_var( section['start'] if 'start' in section.keys() else [0, 0], inp_data)
    header = replace_var( section['header'] if 'header' in section.keys() else True, inp_data)
    border = replace_var( section['border'] if 'border' in section.keys() else False, inp_data)
    columns = section['columns'] if 'columns' in section.keys() else []

    if header:
        for idx, col in enumerate(columns):
            cell = worksheet.cell(row = position[0], column = position[1] + idx)
            cell.value = col['name']

            if border:
                cell.border = Border(left = Side(border_style=None, color='000000', style='thin'),
                                     right = Side(border_style=None, color='000000', style='thin'),
                                     top = Side(border_style=None, color='000000', style='thin'),
                                     bottom = Side(border_style=None, color='000000', style='thin'))

                cell.font = Font(bold = True)

    for row_idx, row in enumerate(table_data, start = 1 if header else 0):
        for col_idx, col in enumerate(columns):
            col_val = replace_var(col['data'], row)
            cell = worksheet.cell(row = position[0] + row_idx, column = position[1] + col_idx)
            cell.value = col_val

            if border:
                cell.border = Border(left = Side(border_style=None, color='000000', style='thin'),
                                     right = Side(border_style=None, color='000000', style='thin'),
                                     top = Side(border_style=None, color='000000', style='thin'),
                                     bottom = Side(border_style=None, color='000000', style='thin'))

    if max_row < position[0] + row_idx:
        max_row = position[0] + row_idx

def replace_var(field, data):
    variable_regex = re.compile(r"\$\{\{([a-z0-9\_\.]{1,})\}\}", re.I)

    if type(field) == type(''):
        template_variables = variable_regex.findall(field)

        for template_variable in template_variables:
            variable_value = ''

            if template_variable == 'max_row':
                variable_value = str(max_row)

            elif template_variable in data.keys():
                variable_value = data[template_variable]

            elif '.' in template_variable:
                subfields = template_variable.split('.')
                subdata = data
                notfound = 0
                for subfield in subfields:
                    if subfield in subdata.keys():
                        subdata = subdata[subfield]
                    else:
                        notfound = 1
                        break

                if notfound == 0:
                    variable_value = subdata

            if type(variable_value) != type('') and field.strip() == f'${{{{{template_variable}}}}}':
                field = variable_value

            else:
                field = field.replace(f'${{{{{template_variable}}}}}', variable_value)
                
        return field

    elif type(field) == type(dict()):
        for fl in field.keys():
            field[fl] = replace_var(field[fl], data)
        
        return field

    elif type(field) == type([]):
        return [replace_var(fl, data) for fl in field]
    
    else:
        return field